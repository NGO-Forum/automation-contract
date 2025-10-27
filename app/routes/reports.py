from flask import Blueprint, render_template, request, send_file
from flask_login import login_required
from app import db
from app.models.contract import Contract
from app.models.department import Department
from app.models.user import User
from datetime import datetime, timedelta
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

reports_bp = Blueprint('reports', __name__)

@reports_bp.route('/contracts')
@login_required
def contract_report():
    # Filters
    department_id = request.args.get('department_id', 'all')
    search = request.args.get('search', '').strip().lower()
    month_year = request.args.get('month_year', datetime.now().strftime('%B %Y'))
    sort = request.args.get('sort', 'contract_number_asc')
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)
    view_mode = request.args.get('view_mode', 'monthly')
    day_filter = request.args.get('day_filter', 'All')

    # Parse month_year
    try:
        month_name, year_str = month_year.split()
        month_num = datetime.strptime(month_name, '%B').month
        year = int(year_str)
        # Convert full month name to abbreviated for display
        month_abbr = datetime.strptime(month_name, '%B').strftime('%b')
        month_year_display = f"{month_abbr} {year}"
    except (ValueError, AttributeError):
        month_num = datetime.now().month
        year = datetime.now().year
        month_year = datetime.now().strftime('%B %Y')
        month_year_display = datetime.now().strftime('%b %Y')

    # Day map for dayofweek
    day_map = {'Mon': 2, 'Tue': 3, 'Wed': 4, 'Thu': 5, 'Fri': 6, 'Sat': 7, 'Sun': 1}

    # Base query for table (filtered by month and year)
    query = Contract.query.filter(Contract.deleted_at == None)\
                          .outerjoin(Contract.user)\
                          .filter(db.extract('year', Contract.created_at) == year)\
                          .filter(db.extract('month', Contract.created_at) == month_num)

    if department_id != 'all':
        query = query.filter(Contract.user.has(department_id=department_id))

    if search:
        query = query.filter(
            (Contract.project_title.ilike(f'%{search}%')) |
            (Contract.party_b_signature_name.ilike(f'%{search}%')) |
            (User.username.ilike(f'%{search}%') & (Contract.user_id != None))
        )

    if day_filter != 'All':
        dow = day_map.get(day_filter)
        if dow:
            query = query.filter(db.func.dayofweek(Contract.created_at) == dow)

    # Sorting
    if sort == 'contract_number_desc':
        query = query.order_by(Contract.contract_number.desc())
    elif sort == 'project_title_asc':
        query = query.order_by(Contract.project_title.asc())
    elif sort == 'project_title_desc':
        query = query.order_by(Contract.project_title.desc())
    else:
        query = query.order_by(Contract.contract_number.asc())

    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    contracts = [contract.to_dict() for contract in pagination.items]

    # Totals
    total_contracts = query.count()
    departments = Department.query.all()
    department_totals = {}
    if department_id == 'all':
        for dept in departments:
            dept_query = Contract.query.filter(Contract.deleted_at == None)\
                                       .outerjoin(Contract.user)\
                                       .filter(Contract.user.has(department_id=dept.id))\
                                       .filter(db.extract('year', Contract.created_at) == year)\
                                       .filter(db.extract('month', Contract.created_at) == month_num)
            if search:
                dept_query = dept_query.filter(
                    (Contract.project_title.ilike(f'%{search}%')) |
                    (Contract.party_b_signature_name.ilike(f'%{search}%')) |
                    (User.username.ilike(f'%{search}%') & (Contract.user_id != None))
                )
            if day_filter != 'All':
                dow = day_map.get(day_filter)
                if dow:
                    dept_query = dept_query.filter(db.func.dayofweek(Contract.created_at) == dow)
            department_totals[dept.name] = dept_query.count()
    else:
        selected_dept = Department.query.get(department_id)
        if selected_dept:
            dept_query = Contract.query.filter(Contract.deleted_at == None)\
                                       .outerjoin(Contract.user)\
                                       .filter(Contract.user.has(department_id=department_id))\
                                       .filter(db.extract('year', Contract.created_at) == year)\
                                       .filter(db.extract('month', Contract.created_at) == month_num)
            if search:
                dept_query = dept_query.filter(
                    (Contract.project_title.ilike(f'%{search}%')) |
                    (Contract.party_b_signature_name.ilike(f'%{search}%')) |
                    (User.username.ilike(f'%{search}%') & (Contract.user_id != None))
                )
            if day_filter != 'All':
                dow = day_map.get(day_filter)
                if dow:
                    dept_query = dept_query.filter(db.func.dayofweek(Contract.created_at) == dow)
            department_totals[selected_dept.name] = dept_query.count()

    # Unique month_years (abbreviated for display)
    try:
        unique_months = db.session.query(
            db.func.distinct(db.func.date_format(Contract.created_at, '%M %Y'))
        ).filter(Contract.created_at != None).all()
        unique_months = [datetime.strptime(m[0], '%B %Y').strftime('%b %Y') for m in unique_months if m[0]]
    except:
        unique_months = [month_year_display]

    # Chart Data
    if view_mode == 'monthly':
        # Monthly counts for the year
        try:
            monthly_query = db.session.query(
                db.func.month(Contract.created_at),
                db.func.count(Contract.id)
            ).filter(
                Contract.deleted_at == None,
                db.extract('year', Contract.created_at) == year
            )
            if department_id != 'all':
                monthly_query = monthly_query.filter(Contract.user.has(department_id=department_id))
            if search:
                monthly_query = monthly_query.outerjoin(Contract.user).filter(
                    (Contract.project_title.ilike(f'%{search}%')) |
                    (Contract.party_b_signature_name.ilike(f'%{search}%')) |
                    (User.username.ilike(f'%{search}%') & (Contract.user_id != None))
                )
            if day_filter != 'All':
                dow = day_map.get(day_filter)
                if dow:
                    monthly_query = monthly_query.filter(db.func.dayofweek(Contract.created_at) == dow)
            monthly_counts = monthly_query.group_by(
                db.func.month(Contract.created_at)
            ).all()

            monthly_dict = {m: c for m, c in monthly_counts}
            chart_labels = [
                'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec'
            ]
            chart_values = [monthly_dict.get(i, 0) for i in range(1, 13)]
            chart_title = f"Contracts Overview (Jan - Dec {year})"
        except:
            chart_labels = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
            chart_values = [0] * 12
            chart_title = f"Contracts Overview (Jan - Dec {year})"
    else:
        # Weekly view: Aggregated by day of week (Monday to Sunday) for the selected month
        try:
            first_day = datetime(year, month_num, 1)
            last_day = (first_day + timedelta(days=31)).replace(day=1) - timedelta(days=1)

            day_query = db.session.query(
                db.func.dayofweek(Contract.created_at),
                db.func.count(Contract.id)
            ).filter(
                Contract.deleted_at == None,
                Contract.created_at >= first_day,
                Contract.created_at < first_day + timedelta(days=32 - first_day.day)
            )
            if department_id != 'all':
                day_query = day_query.filter(Contract.user.has(department_id=department_id))
            if search:
                day_query = day_query.outerjoin(Contract.user).filter(
                    (Contract.project_title.ilike(f'%{search}%')) |
                    (Contract.party_b_signature_name.ilike(f'%{search}%')) |
                    (User.username.ilike(f'%{search}%') & (Contract.user_id != None))
                )
            if day_filter != 'All':
                dow = day_map.get(day_filter)
                if dow:
                    day_query = day_query.filter(db.func.dayofweek(Contract.created_at) == dow)
            day_counts = day_query.group_by(
                db.func.dayofweek(Contract.created_at)
            ).all()

            # DAYOFWEEK: 1=Sunday, 2=Monday, ..., 7=Saturday
            dow_dict = {int(dow): count for dow, count in day_counts}

            # Map to Monday (2) to Sunday (1)
            chart_labels = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
            chart_values = [
                dow_dict.get(2, 0),  # Monday
                dow_dict.get(3, 0),  # Tuesday
                dow_dict.get(4, 0),  # Wednesday
                dow_dict.get(5, 0),  # Thursday
                dow_dict.get(6, 0),  # Friday
                dow_dict.get(7, 0),  # Saturday
                dow_dict.get(1, 0)   # Sunday
            ]

            # Ensure at least one data point
            if sum(chart_values) == 0:
                chart_labels = ['No Data']
                chart_values = [0]
            chart_title = f"Contracts by Day of Week ({month_year_display})"
        except:
            chart_labels = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
            chart_values = [0] * 7
            chart_title = f"Contracts by Day of Week ({month_year_display})"

    # Pie chart (department distribution for the month)
    pie_labels = list(department_totals.keys())
    pie_values = list(department_totals.values())
    pie_title = "Contracts by Department"
    if day_filter != 'All':
        pie_title += f" (on {day_filter}s)"
    if department_id != 'all':
        selected_dept = Department.query.get(department_id)
        if selected_dept:
            pie_title += f" ({selected_dept.name})"

    return render_template('reports/index.html',
                           contracts=contracts,
                           pagination=pagination,
                           departments=departments,
                           department_id=department_id,
                           search=search,
                           month_year=month_year,  # Keep full month name for URL/filter consistency
                           month_year_display=month_year_display,  # Abbreviated for display
                           sort=sort,
                           per_page=per_page,
                           total_contracts=total_contracts,
                           department_totals=department_totals,
                           unique_months=unique_months,
                           chart_labels=chart_labels,
                           chart_values=chart_values,
                           pie_labels=pie_labels,
                           pie_values=pie_values,
                           chart_title=chart_title,
                           pie_title=pie_title,
                           view_mode=view_mode,
                           day_filter=day_filter)

def safe_create_workbook():
    """Safely create a workbook with basic formatting"""
    try:
        wb = Workbook()
        # Remove default sheet
        if wb.worksheets:
            wb.remove(wb.active)
        return wb
    except:
        # Fallback: create minimal workbook
        from openpyxl import Workbook as FallbackWorkbook
        return FallbackWorkbook()

def safe_add_sheet(wb, title):
    """Safely add a sheet with the given title"""
    try:
        # Clean sheet name (Excel limits)
        safe_title = str(title)[:31].strip()
        if not safe_title:
            safe_title = "Sheet1"
        
        # Check if sheet already exists
        for sheet in wb.worksheets:
            if sheet.title == safe_title:
                return sheet
        
        return wb.create_sheet(safe_title)
    except:
        return wb.create_sheet("Data")

def safe_write_cell(ws, row, col, value):
    """Safely write a value to a cell"""
    try:
        if ws and row > 0 and col > 0:
            cell = ws.cell(row=row, column=col)
            cell.value = str(value) if value is not None else ""
            return True
        return False
    except:
        return False

def safe_merge_cells(ws, range_str):
    """Safely merge cells"""
    try:
        if ws and range_str:
            ws.merge_cells(range_str)
            return True
        return False
    except:
        return False

def safe_set_cell_style(cell, font=None, alignment=None, border=None, fill=None):
    """Safely apply styles to a cell"""
    try:
        if cell:
            if font:
                cell.font = font
            if alignment:
                cell.alignment = alignment
            if border:
                cell.border = border
            if fill:
                cell.fill = fill
            return True
        return False
    except:
        return False

def safe_set_column_width(ws, col_letter, width):
    """Safely set column width"""
    try:
        if ws and col_letter:
            ws.column_dimensions[col_letter].width = min(float(width), 150)
            return True
        return False
    except:
        return False

def safe_set_row_height(ws, row_num, height):
    """Safely set row height"""
    try:
        if ws and row_num > 0:
            ws.row_dimensions[row_num].height = float(height)
            return True
        return False
    except:
        return False

def safe_apply_styling(ws, title_row, filter_row, data_start_row, total_row, num_cols=6):
    """Safely apply styling to worksheet with custom colors 1E659E for title"""
    try:
        if not ws:
            return False
        
        # Define styles with updated font sizes and custom colors
        title_font = Font(bold=True, name='Calibri', size=18, color="FFFFFF")  # Title: 18px white text
        filter_font = Font(name='Calibri', size=12, italic=True, color="555555")  # Filter: 12px
        header_font = Font(bold=True, name='Calibri', size=16, color="FFFFFF")  # Headers: 16px white text
        data_font = Font(name='Calibri', size=12, color="000000")  # Data rows: 12px
        total_font = Font(bold=True, name='Calibri', size=14, color="000000")  # Total: 14px
        
        # Increased padding alignments
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        wrap_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=2)  # Increased indent
        
        thin_border = Border(
            left=Side(style='thin', color="1E659E"),  # Custom blue border color
            right=Side(style='thin', color="1E659E"),
            top=Side(style='thin', color="1E659E"),
            bottom=Side(style='thin', color="1E659E")
        )
        
        # Custom blue color #1E659E for title, green #93A843 for headers
        title_fill = PatternFill(start_color='1E659E', end_color='1E659E', fill_type='solid')  # Custom blue for title
        header_fill = PatternFill(start_color='93A843', end_color='93A843', fill_type='solid')  # Custom green for headers
        total_fill = PatternFill(start_color='E8F5E8', end_color='E8F5E8', fill_type='solid')  # Light green for total
        data_alternate_fill = PatternFill(start_color='F9FFF9', end_color='F9FFF9', fill_type='solid')  # Very subtle green for alternating rows

        # Style title row with custom blue background and white text
        if title_row > 0:
            for col in range(1, num_cols + 1):
                cell = ws.cell(row=title_row, column=col)
                safe_set_cell_style(cell, title_font, center_alignment, thin_border, title_fill)

        # Style filter row if exists
        if filter_row > 0 and filter_row != title_row:
            for col in range(1, num_cols + 1):
                cell = ws.cell(row=filter_row, column=col)
                safe_set_cell_style(cell, filter_font, center_alignment, thin_border)

        # Style headers (Column titles) - 16px with custom green background and white text
        if data_start_row > 0:
            for col in range(1, num_cols + 1):
                cell = ws.cell(row=data_start_row, column=col)
                safe_set_cell_style(cell, header_font, center_alignment, thin_border, header_fill)

        # Style data rows - 12px for all data cells with increased spacing and alternating colors
        if data_start_row > 0:
            data_end_row = total_row - 1 if total_row > 0 else ws.max_row
            for row_num in range(data_start_row + 1, data_end_row + 1):
                # Alternate row colors for better readability
                use_alternate_color = (row_num - data_start_row - 1) % 2 == 1
                row_fill = data_alternate_fill if use_alternate_color else None
                
                for col in range(1, num_cols + 1):
                    cell = ws.cell(row=row_num, column=col)
                    if col == 2:  # Project Title column - left aligned with wrap
                        safe_set_cell_style(cell, data_font, wrap_alignment, thin_border, row_fill)
                    else:
                        alignment = left_alignment if col in [3, 4] else center_alignment
                        safe_set_cell_style(cell, data_font, alignment, thin_border, row_fill)

        # Style total row - ensure it exists and has proper styling
        if total_row > 0 and total_row <= ws.max_row + 1:
            # Apply border and fill to all columns in total row
            for col in range(1, num_cols + 1):
                cell = ws.cell(row=total_row, column=col)
                safe_set_cell_style(cell, total_font, center_alignment, thin_border, total_fill)
            
            # Special styling for merged cells A and B
            if safe_write_cell(ws, total_row, 1, 'Total Contracts'):
                cell_a = ws.cell(row=total_row, column=1)
                safe_set_cell_style(cell_a, total_font, left_alignment, thin_border, total_fill)
            
            if safe_write_cell(ws, total_row, 2, len([row for row in ws.iter_rows(min_row=data_start_row + 1, max_row=total_row - 1, values_only=True) if row])):
                cell_b = ws.cell(row=total_row, column=2)
                safe_set_cell_style(cell_b, total_font, center_alignment, thin_border, total_fill)

        # Set INCREASED row heights for better spacing
        if title_row > 0:
            safe_set_row_height(ws, title_row, 40)  # Significantly increased for 18px font
        if filter_row > 0 and filter_row != title_row:
            safe_set_row_height(ws, filter_row, 30)  # Increased for 12px font
        if data_start_row > 0:
            safe_set_row_height(ws, data_start_row, 35)  # Increased for 16px header font
        if total_row > 0:
            safe_set_row_height(ws, total_row, 35)  # Increased for 14px font
        
        # Set default row height for data rows to accommodate 12px font with extra space
        try:
            for row_num in range(data_start_row + 1, total_row):
                safe_set_row_height(ws, row_num, 30)  # Increased from 20 to 30
        except:
            pass

        # Freeze panes below headers
        try:
            if data_start_row > 0:
                ws.freeze_panes = ws.cell(row=data_start_row + 1, column=1).coordinate
        except:
            pass

        return True
    except Exception as e:
        print(f"Styling error: {e}")
        return False

def safe_auto_size_columns(ws, num_cols=6):
    """Completely safe column auto-sizing with significantly increased widths"""
    try:
        if not ws:
            return False
        
        # Significantly increased default widths for better spacing
        default_widths = {
            'A': 25,   # Number of Contract - increased for better spacing
            'B': 70,   # Project Title (much wider for long titles and 12px font)
            'C': 30,   # Department - increased for better readability
            'D': 25,   # Manager - increased for better spacing
            'E': 30,   # Contract Date - increased for better readability
            'F': 35    # Party B (much wider for names and 12px font)
        }
        
        for i in range(1, num_cols + 1):
            col_letter = chr(64 + i)
            # Use default width
            safe_set_column_width(ws, col_letter, default_widths.get(col_letter, 30))
        
        return True
    except:
        # Fallback: set all columns to default width
        try:
            for col_letter in ['A', 'B', 'C', 'D', 'E', 'F']:
                safe_set_column_width(ws, col_letter, 30)
        except:
            pass
        return False

def safe_prepare_contract_data(contracts):
    """Safely prepare contract data for export"""
    try:
        if not contracts:
            return []
        
        data = []
        for c in contracts:
            try:
                dept_name = 'N/A'
                manager_name = 'N/A'
                if hasattr(c, 'user') and c.user:
                    if hasattr(c.user, 'department') and c.user.department:
                        dept_name = getattr(c.user.department, 'name', 'N/A') or 'N/A'
                    if hasattr(c.user, 'username'):
                        manager_name = c.user.username or 'N/A'
                
                data.append({
                    'Number of Contract': getattr(c, 'contract_number', 'N/A') or 'N/A',
                    'Project Title': getattr(c, 'project_title', 'N/A') or 'N/A',
                    'Department': dept_name,
                    'Manager': manager_name,
                    'Contract Date': getattr(c, 'formatted_created_at', str(c.created_at)[:10] if hasattr(c, 'created_at') and c.created_at else 'N/A'),
                    'Party B': getattr(c, 'party_b_signature_name', 'N/A') or 'N/A'
                })
            except Exception as e:
                print(f"Error preparing contract data: {e}")
                # Skip problematic contracts
                continue
        
        return data
    except Exception as e:
        print(f"Error in prepare_contract_data: {e}")
        return []

def safe_export_single_sheet(contracts, department_name="", month_year="", search="", day_filter="All"):
    """Create a single safe export sheet with improved formatting and custom colors"""
    try:
        print(f"Exporting {len(contracts)} contracts for {department_name}")
        
        wb = safe_create_workbook()
        ws = safe_add_sheet(wb, "Contract Report")
        
        # Title row with extra spacing - now uses #1E659E
        title_row = 1
        title_text = f"CONTRACT REPORT - {month_year}"
        if department_name:
            title_text += f" ({department_name})"
        safe_write_cell(ws, title_row, 1, title_text)
        safe_merge_cells(ws, f'A{title_row}:F{title_row}')
        
        # Add extra blank row for spacing after title
        blank_row_1 = title_row + 1
        for col in range(1, 7):
            safe_write_cell(ws, blank_row_1, col, "")
        
        # Filter info row
        filter_row = title_row + 3  # Increased spacing
        filter_info = []
        if department_name and not department_name in title_text:
            filter_info.append(f"Department: {department_name}")
        if search:
            filter_info.append(f"Search: {search}")
        if day_filter != "All":
            filter_info.append(f"Day Filter: {day_filter}")
        
        has_filter_row = False
        if filter_info:
            filter_text = " | ".join(filter_info)
            safe_write_cell(ws, filter_row, 1, filter_text)
            safe_merge_cells(ws, f'A{filter_row}:F{filter_row}')
            has_filter_row = True
            data_start_row = filter_row + 2  # Extra space after filter
        else:
            data_start_row = filter_row + 1
        
        # Add another blank row before headers for extra spacing
        blank_row_2 = data_start_row
        for col in range(1, 7):
            safe_write_cell(ws, blank_row_2, col, "")
        data_start_row += 1  # Move headers down one more row
        
        # Column headers (16px font with custom green #93A843)
        headers = ['Number of Contract', 'Project Title', 'Department', 'Manager', 'Contract Date', 'Party B']
        for col, header in enumerate(headers, 1):
            safe_write_cell(ws, data_start_row, col, header)
        
        # Data rows (12px font)
        data = safe_prepare_contract_data(contracts)
        print(f"Prepared {len(data)} data rows")
        
        current_row = data_start_row + 1
        for row_data in data:
            for col, value in enumerate(row_data.values(), 1):
                safe_write_cell(ws, current_row, col, value)
            current_row += 1
        
        # Calculate total row position
        total_row = current_row + 1  # Extra space before total row
        print(f"Total row position: {total_row}")
        
        # Write Total row - ensure it's written properly
        total_contracts_count = len(data)
        print(f"Total contracts count: {total_contracts_count}")
        
        # Clear the total row first
        for col in range(1, 7):
            ws.cell(row=total_row, column=col).value = None
        
        # Write Total Contracts text
        safe_write_cell(ws, total_row, 1, 'Total Contracts')
        
        # Write the count in column B
        safe_write_cell(ws, total_row, 2, total_contracts_count)
        
        # Merge cells A and B for the total row
        safe_merge_cells(ws, f'A{total_row}:B{total_row}')
        
        print(f"Total row written: A{total_row}:B{total_row} with value {total_contracts_count}")
        
        # Apply styling AFTER all data is written
        styling_success = safe_apply_styling(ws, title_row, filter_row if has_filter_row else 0, data_start_row, total_row)
        print(f"Styling success: {styling_success}")
        
        # Auto-size columns with increased widths
        safe_auto_size_columns(ws)
        
        # Set active sheet
        try:
            wb.active = ws
        except:
            pass
        
        # Save
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"Contract_Report_{month_year.replace(' ', '_')}"
        if department_name:
            filename += f"_{department_name.replace(' ', '_')}"
        filename += ".xlsx"
        
        return output, filename
        
    except Exception as e:
        print(f"Error in safe_export_single_sheet: {e}")
        # Ultimate fallback - simple export with custom colors
        try:
            print("Using fallback export")
            wb = Workbook()
            ws = wb.active
            ws.title = "Contract Report"
            
            # Custom color styles
            header_font = Font(bold=True, name='Calibri', size=16, color="FFFFFF")  # White text for green header
            data_font = Font(name='Calibri', size=12, color="000000")
            title_font = Font(bold=True, name='Calibri', size=18, color="FFFFFF")  # White text for blue title
            total_font = Font(bold=True, name='Calibri', size=14, color="000000")
            
            title_fill = PatternFill(start_color='1E659E', end_color='1E659E', fill_type='solid')  # Custom blue for title
            header_fill = PatternFill(start_color='93A843', end_color='93A843', fill_type='solid')  # Custom green for headers
            total_fill = PatternFill(start_color='E8F5E8', end_color='E8F5E8', fill_type='solid')  # Light green
            data_alternate_fill = PatternFill(start_color='F9FFF9', end_color='F9FFF9', fill_type='solid')  # Subtle green
            
            # Title with spacing and custom blue background
            ws.append(["CONTRACT REPORT"])
            ws.row_dimensions[1].height = 40
            for cell in ws[1]:
                cell.font = title_font
                cell.alignment = Alignment(horizontal='center')
                cell.fill = title_fill
            
            # Blank row for spacing
            ws.append([""])
            ws.row_dimensions[2].height = 15
            
            # Headers with custom green background
            headers = ['Number of Contract', 'Project Title', 'Department', 'Manager', 'Contract Date', 'Party B']
            header_row = ws.append(headers)
            ws.row_dimensions[3].height = 35
            for cell in ws[header_row]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            # Simple data (limit to 100 records) with 12px font and alternating colors
            export_count = 0
            for i, contract in enumerate(contracts[:100]):
                try:
                    row_data = [
                        getattr(contract, 'contract_number', 'N/A') or 'N/A',
                        getattr(contract, 'project_title', 'N/A') or 'N/A',
                        'N/A',  # Department
                        'N/A',  # Manager
                        getattr(contract, 'formatted_created_at', 'N/A') or 'N/A',
                        getattr(contract, 'party_b_signature_name', 'N/A') or 'N/A'
                    ]
                    data_row = ws.append(row_data)
                    ws.row_dimensions[data_row].height = 30
                    
                    # Alternate row colors
                    if i % 2 == 1:
                        for cell in ws[data_row]:
                            cell.font = data_font
                            cell.fill = data_alternate_fill
                    else:
                        for cell in ws[data_row]:
                            cell.font = data_font
                    
                    export_count += 1
                except Exception as row_error:
                    print(f"Row error: {row_error}")
                    continue
            
            # Blank row before total
            ws.append([""])
            ws.row_dimensions[len([row for row in ws.iter_rows()])].height = 10
            
            # Add total row in fallback with 14px font
            total_row_fallback = len([row for row in ws.iter_rows(values_only=True)]) + 1
            ws.cell(row=total_row_fallback, column=1, value='Total Contracts').font = total_font
            ws.cell(row=total_row_fallback, column=2, value=export_count).font = total_font
            ws.merge_cells(start_row=total_row_fallback, start_column=1, end_row=total_row_fallback, end_column=2)
            ws.row_dimensions[total_row_fallback].height = 35
            
            # Apply total row fill
            for col in range(1, 7):
                cell = ws.cell(row=total_row_fallback, column=col)
                cell.fill = total_fill
            
            # Set column widths
            for col_letter, width in [('A', 25), ('B', 70), ('C', 30), ('D', 25), ('E', 30), ('F', 35)]:
                ws.column_dimensions[col_letter].width = width
            
            # Add borders with custom blue color
            thin_border = Border(
                left=Side(style='thin', color="1E659E"),
                right=Side(style='thin', color="1E659E"),
                top=Side(style='thin', color="1E659E"),
                bottom=Side(style='thin', color="1E659E")
            )
            
            # Apply borders to header row
            for cell in ws[3]:
                cell.border = thin_border
            
            # Apply borders to data rows
            for row_num in range(4, total_row_fallback):
                for cell in ws[row_num]:
                    cell.border = thin_border
            
            # Apply borders to total row
            for cell in ws[total_row_fallback]:
                cell.border = thin_border
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return output, "Contract_Report_Fallback.xlsx"
        except Exception as fallback_error:
            print(f"Fallback error: {fallback_error}")
            # Absolute last resort - empty file
            output = BytesIO()
            wb = Workbook()
            ws = wb.active
            header_font = Font(bold=True, name='Calibri', size=16, color="FFFFFF")  # White text for green header
            data_font = Font(name='Calibri', size=12, color="000000")
            total_font = Font(bold=True, name='Calibri', size=14, color="000000")
            title_font = Font(bold=True, name='Calibri', size=18, color="FFFFFF")  # White text for blue title
            
            title_fill = PatternFill(start_color='1E659E', end_color='1E659E', fill_type='solid')  # Custom blue for title
            header_fill = PatternFill(start_color='93A843', end_color='93A843', fill_type='solid')  # Custom green
            total_fill = PatternFill(start_color='E8F5E8', end_color='E8F5E8', fill_type='solid')  # Light green
            
            # Title
            ws.append(["CONTRACT REPORT"])
            ws.row_dimensions[1].height = 40
            for cell in ws[1]:
                cell.font = title_font
                cell.alignment = Alignment(horizontal='center')
                cell.fill = title_fill
            
            # Blank row
            ws.append([""])
            ws.row_dimensions[2].height = 15
            
            # Headers
            headers = ['Number of Contract', 'Project Title', 'Department', 'Manager', 'Contract Date', 'Party B']
            header_row = ws.append(headers)
            ws.row_dimensions[3].height = 35
            for cell in ws[header_row]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            # Total row
            ws.append([""])
            ws.row_dimensions[4].height = 10
            total_row = 5
            ws.cell(row=total_row, column=1, value='Total Contracts').font = total_font
            ws.cell(row=total_row, column=2, value=0).font = total_font
            ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=2)
            ws.row_dimensions[total_row].height = 35
            
            # Apply total row fill
            for col in range(1, 7):
                cell = ws.cell(row=total_row, column=col)
                cell.fill = total_fill
            
            # Set column widths
            for col_letter, width in [('A', 25), ('B', 70), ('C', 30), ('D', 25), ('E', 30), ('F', 35)]:
                ws.column_dimensions[col_letter].width = width
            
            # Add borders with custom blue color
            thin_border = Border(
                left=Side(style='thin', color="1E659E"),
                right=Side(style='thin', color="1E659E"),
                top=Side(style='thin', color="1E659E"),
                bottom=Side(style='thin', color="1E659E")
            )
            
            # Apply borders
            for row_num in range(1, total_row + 1):
                for cell in ws[row_num]:
                    cell.border = thin_border
            
            wb.save(output)
            output.seek(0)
            return output, "Contract_Report_Empty.xlsx"

def safe_export_all_departments(departments, year, month_num, month_year, search="", day_filter="All"):
    """Safely create multi-sheet export for all departments with proper filters and custom colors"""
    try:
        print(f"Exporting all departments for {month_year}")
        wb = safe_create_workbook()
        
        # Day map for dayofweek
        day_map = {'Mon': 2, 'Tue': 3, 'Wed': 4, 'Thu': 5, 'Fri': 6, 'Sat': 7, 'Sun': 1}
        dow = day_map.get(day_filter) if day_filter != 'All' else None
        
        # Helper to get filtered contracts
        def get_filtered_contracts(dept_id=None):
            try:
                q = Contract.query.filter(Contract.deleted_at == None)\
                                  .outerjoin(Contract.user)\
                                  .filter(db.extract('year', Contract.created_at) == year)\
                                  .filter(db.extract('month', Contract.created_at) == month_num)

                if dept_id and dept_id != 'all':
                    q = q.filter(Contract.user.has(department_id=dept_id))

                if search:
                    q = q.filter(
                        (Contract.project_title.ilike(f'%{search}%')) |
                        (Contract.party_b_signature_name.ilike(f'%{search}%')) |
                        (User.username.ilike(f'%{search}%') & (Contract.user_id != None))
                    )

                if dow:
                    q = q.filter(db.func.dayofweek(Contract.created_at) == dow)

                q = q.order_by(Contract.contract_number.asc())
                return q.all()
            except Exception as e:
                print(f"Error getting filtered contracts: {e}")
                return []
        
        # All contracts for main sheet
        all_contracts = get_filtered_contracts()
        print(f"Found {len(all_contracts)} total contracts")
        
        # Create main sheet for all departments
        main_ws = safe_add_sheet(wb, "All Departments")
        
        # Write main sheet data with extra spacing
        title_row = 1
        title_text = f"CONTRACT REPORT - {month_year} (All Departments)"
        safe_write_cell(main_ws, title_row, 1, title_text)
        safe_merge_cells(main_ws, f'A{title_row}:F{title_row}')
        
        # Blank row after title
        blank_row_1 = title_row + 1
        for col in range(1, 7):
            safe_write_cell(main_ws, blank_row_1, col, "")
        
        filter_row = title_row + 3  # Increased spacing
        filter_info = []
        if search:
            filter_info.append(f"Search: {search}")
        if day_filter != "All":
            filter_info.append(f"Day Filter: {day_filter}")
        
        has_filter_row = False
        if filter_info:
            filter_text = " | ".join(filter_info)
            safe_write_cell(main_ws, filter_row, 1, filter_text)
            safe_merge_cells(main_ws, f'A{filter_row}:F{filter_row}')
            has_filter_row = True
            data_start_row = filter_row + 2  # Extra space after filter
        else:
            data_start_row = filter_row + 1
        
        # Blank row before headers
        blank_row_2 = data_start_row
        for col in range(1, 7):
            safe_write_cell(main_ws, blank_row_2, col, "")
        data_start_row += 1
        
        headers = ['Number of Contract', 'Project Title', 'Department', 'Manager', 'Contract Date', 'Party B']
        for col, header in enumerate(headers, 1):
            safe_write_cell(main_ws, data_start_row, col, header)
        
        main_data = safe_prepare_contract_data(all_contracts)
        current_row = data_start_row + 1
        for row_data in main_data:
            for col, value in enumerate(row_data.values(), 1):
                safe_write_cell(main_ws, current_row, col, value)
            current_row += 1
        
        # Total row for main sheet with extra spacing
        total_row = current_row + 1  # Extra space before total
        safe_write_cell(main_ws, total_row, 1, 'Total Contracts')
        safe_write_cell(main_ws, total_row, 2, len(all_contracts))
        safe_merge_cells(main_ws, f'A{total_row}:B{total_row}')
        
        safe_apply_styling(main_ws, title_row, filter_row if has_filter_row else 0, data_start_row, total_row)
        safe_auto_size_columns(main_ws)
        
        # Individual department sheets with increased spacing and custom colors
        dept_count = 0
        for dept in departments:
            try:
                dept_contracts = get_filtered_contracts(dept.id)
                if not dept_contracts:
                    continue
                
                dept_count += 1
                sheet_name = str(dept.name)[:31].strip()
                if not sheet_name:
                    sheet_name = f"Dept_{dept.id}"
                
                ws_dept = safe_add_sheet(wb, sheet_name)
                
                # Department title with spacing - uses #1E659E
                title_row = 1
                title_text = f"{dept.name} - CONTRACT REPORT - {month_year}"
                safe_write_cell(ws_dept, title_row, 1, title_text)
                safe_merge_cells(ws_dept, f'A{title_row}:F{title_row}')
                
                # Blank row after title
                blank_row_1 = title_row + 1
                for col in range(1, 7):
                    safe_write_cell(ws_dept, blank_row_1, col, "")
                
                filter_row = title_row + 3
                has_dept_filter = False
                if filter_info:
                    safe_write_cell(ws_dept, filter_row, 1, filter_text)
                    safe_merge_cells(ws_dept, f'A{filter_row}:F{filter_row}')
                    has_dept_filter = True
                    data_start_row = filter_row + 2
                else:
                    data_start_row = filter_row + 1
                
                # Blank row before headers
                blank_row_2 = data_start_row
                for col in range(1, 7):
                    safe_write_cell(ws_dept, blank_row_2, col, "")
                data_start_row += 1
                
                # Headers (16px with custom green #93A843)
                for col, header in enumerate(headers, 1):
                    safe_write_cell(ws_dept, data_start_row, col, header)
                
                # Data (12px with alternating colors)
                dept_data = safe_prepare_contract_data(dept_contracts)
                current_row = data_start_row + 1
                for row_data in dept_data:
                    for col, value in enumerate(row_data.values(), 1):
                        safe_write_cell(ws_dept, current_row, col, value)
                    current_row += 1
                
                # Department total row with extra spacing
                total_row = current_row + 1
                safe_write_cell(ws_dept, total_row, 1, f'{dept.name} Total')
                safe_write_cell(ws_dept, total_row, 2, len(dept_contracts))
                safe_merge_cells(ws_dept, f'A{total_row}:B{total_row}')
                
                safe_apply_styling(ws_dept, title_row, filter_row if has_dept_filter else 0, data_start_row, total_row)
                safe_auto_size_columns(ws_dept)
                
            except Exception as dept_error:
                print(f"Error creating department sheet for {dept.name}: {dept_error}")
                continue
        
        print(f"Created {dept_count} department sheets")
        
        # Set active sheet to main
        try:
            wb.active = wb["All Departments"]
        except:
            pass
        
        # Save
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output, f"All_Departments_Report_{month_year.replace(' ', '_')}.xlsx"
        
    except Exception as e:
        print(f"Error in safe_export_all_departments: {e}")
        # Fallback to single sheet with all contracts
        all_contracts = get_filtered_contracts()
        return safe_export_single_sheet(all_contracts, "All Departments", month_year, search, day_filter)

@reports_bp.route('/export_contracts_excel')
@login_required
def export_contracts_excel():
    department_id = request.args.get('department_id', 'all')
    month_year = request.args.get('month_year', datetime.now().strftime('%B %Y'))
    search = request.args.get('search', '').strip().lower()
    day_filter = request.args.get('day_filter', 'All')

    try:
        print(f"Export request: dept={department_id}, month={month_year}, search={search}, day={day_filter}")
        
        # Parse month/year for filtering
        try:
            month_name, year_str = month_year.split()
            month_num = datetime.strptime(month_name, '%B').month
            year = int(year_str)
            print(f"Parsed: year={year}, month={month_num}")
        except Exception as parse_error:
            print(f"Parse error: {parse_error}")
            month_num = datetime.now().month
            year = datetime.now().year

        if department_id == 'all':
            # Export all departments
            departments = Department.query.all()
            print(f"Found {len(departments)} departments")
            output, filename = safe_export_all_departments(departments, year, month_num, month_year, search, day_filter)
        else:
            # Export single department
            # Helper function to get contracts
            def get_filtered_contracts(dept_id=None):
                try:
                    q = Contract.query.filter(Contract.deleted_at == None)\
                                      .outerjoin(Contract.user)\
                                      .filter(db.extract('year', Contract.created_at) == year)\
                                      .filter(db.extract('month', Contract.created_at) == month_num)

                    if dept_id and dept_id != 'all':
                        q = q.filter(Contract.user.has(department_id=dept_id))

                    if search:
                        q = q.filter(
                            (Contract.project_title.ilike(f'%{search}%')) |
                            (Contract.party_b_signature_name.ilike(f'%{search}%')) |
                            (User.username.ilike(f'%{search}%') & (Contract.user_id != None))
                        )

                    day_map = {'Mon': 2, 'Tue': 3, 'Wed': 4, 'Thu': 5, 'Fri': 6, 'Sat': 7, 'Sun': 1}
                    if day_filter != 'All':
                        dow = day_map.get(day_filter)
                        if dow:
                            q = q.filter(db.func.dayofweek(Contract.created_at) == dow)

                    # Simple sorting
                    q = q.order_by(Contract.contract_number.asc())
                    return q.all()
                except Exception as q_error:
                    print(f"Query error: {q_error}")
                    return []
            
            contracts = get_filtered_contracts(department_id)
            print(f"Found {len(contracts)} contracts for department {department_id}")
            
            selected_dept = Department.query.get(department_id) if department_id != 'current' else None
            dept_name = selected_dept.name if selected_dept else "Selected Department"
            output, filename = safe_export_single_sheet(contracts, dept_name, month_year, search, day_filter)

        # Ensure filename is safe
        safe_filename = filename.replace('/', '_').replace('\\', '_').replace(':', '_')[:200]
        
        print(f"Sending file: {safe_filename}")
        return send_file(
            output,
            download_name=safe_filename,
            as_attachment=True,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f"Critical export error: {e}")
        # Ultimate fallback - simple CSV-like export with custom colors
        try:
            print("Using ultimate fallback")
            output = BytesIO()
            wb = Workbook()
            ws = wb.active
            ws.title = "Contracts"
            
            # Custom color styles
            header_font = Font(bold=True, name='Calibri', size=16, color="FFFFFF")  # White text for green header
            data_font = Font(name='Calibri', size=12, color="000000")
            total_font = Font(bold=True, name='Calibri', size=14, color="000000")
            title_font = Font(bold=True, name='Calibri', size=18, color="FFFFFF")  # White text for blue title
            
            title_fill = PatternFill(start_color='1E659E', end_color='1E659E', fill_type='solid')  # Custom blue for title
            header_fill = PatternFill(start_color='93A843', end_color='93A843', fill_type='solid')  # Custom green for headers
            total_fill = PatternFill(start_color='E8F5E8', end_color='E8F5E8', fill_type='solid')  # Light green
            data_alternate_fill = PatternFill(start_color='F9FFF9', end_color='F9FFF9', fill_type='solid')  # Subtle green
            
            # Title
            ws.append(["CONTRACTS REPORT"])
            ws.row_dimensions[1].height = 40
            for cell in ws[1]:
                cell.font = title_font
                cell.alignment = Alignment(horizontal='center')
                cell.fill = title_fill
            
            # Blank row
            ws.append([""])
            ws.row_dimensions[2].height = 15
            
            # Headers
            ws.append(['Contract Number', 'Project Title', 'Department', 'Manager', 'Date', 'Party B'])
            ws.row_dimensions[3].height = 35
            for cell in ws[3]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                
            # Get some contracts
            contracts = Contract.query.filter(Contract.deleted_at == None).limit(100).all()
            export_count = 0
            for i, c in enumerate(contracts):
                try:
                    row_data = [
                        c.contract_number or '',
                        c.project_title or '',
                        c.user.department.name if c.user and c.user.department else '',
                        c.user.username if c.user else '',
                        str(c.created_at)[:10] if c.created_at else '',
                        c.party_b_signature_name or ''
                    ]
                    data_row = ws.append(row_data)
                    ws.row_dimensions[data_row].height = 30
                    
                    # Alternate row colors
                    if i % 2 == 1:
                        for cell in ws[data_row]:
                            cell.font = data_font
                            cell.fill = data_alternate_fill
                    else:
                        for cell in ws[data_row]:
                            cell.font = data_font
                    
                    export_count += 1
                except Exception as row_error:
                    print(f"Row processing error: {row_error}")
                    continue
            
            # Blank row before total
            ws.append([""])
            ws.row_dimensions[len([row for row in ws.iter_rows()])].height = 10
            
            # Add total row with 14px font
            total_row = len([row for row in ws.iter_rows(values_only=True)]) + 1
            ws.cell(row=total_row, column=1, value='Total Contracts').font = total_font
            ws.cell(row=total_row, column=2, value=export_count).font = total_font
            ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=2)
            ws.row_dimensions[total_row].height = 35
            
            # Apply total row fill
            for col in range(1, 7):
                cell = ws.cell(row=total_row, column=col)
                cell.fill = total_fill
            
            # Set increased column widths
            for col_letter, width in [('A', 25), ('B', 70), ('C', 30), ('D', 25), ('E', 30), ('F', 35)]:
                ws.column_dimensions[col_letter].width = width
            
            # Add borders with custom blue color
            thin_border = Border(
                left=Side(style='thin', color="1E659E"),
                right=Side(style='thin', color="1E659E"),
                top=Side(style='thin', color="1E659E"),
                bottom=Side(style='thin', color="1E659E")
            )
            
            # Apply borders to all rows
            for row_num in range(1, total_row + 1):
                for cell in ws[row_num]:
                    cell.border = thin_border
            
            wb.save(output)
            output.seek(0)
            return send_file(
                output,
                download_name="contracts_fallback.xlsx",
                as_attachment=True,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        except Exception as fallback_error:
            print(f"Even fallback failed: {fallback_error}")
            # Absolute last resort - return empty file
            output = BytesIO()
            wb = Workbook()
            ws = wb.active
            header_font = Font(bold=True, name='Calibri', size=16, color="FFFFFF")  # White text for green header
            data_font = Font(name='Calibri', size=12, color="000000")
            total_font = Font(bold=True, name='Calibri', size=14, color="000000")
            title_font = Font(bold=True, name='Calibri', size=18, color="FFFFFF")  # White text for blue title
            
            title_fill = PatternFill(start_color='1E659E', end_color='1E659E', fill_type='solid')  # Custom blue for title
            header_fill = PatternFill(start_color='93A843', end_color='93A843', fill_type='solid')  # Custom green
            total_fill = PatternFill(start_color='E8F5E8', end_color='E8F5E8', fill_type='solid')  # Light green
            
            # Title
            ws.append(["CONTRACTS REPORT"])
            ws.row_dimensions[1].height = 40
            for cell in ws[1]:
                cell.font = title_font
                cell.alignment = Alignment(horizontal='center')
                cell.fill = title_fill
            
            # Blank row
            ws.append([""])
            ws.row_dimensions[2].height = 15
            
            # Headers
            ws.append(['Contract Number', 'Project Title', 'Department', 'Manager', 'Date', 'Party B'])
            ws.row_dimensions[3].height = 35
            for cell in ws[3]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            # Blank row
            ws.append([""])
            ws.row_dimensions[4].height = 10
            
            # Total row
            total_row = 5
            ws.cell(row=total_row, column=1, value='Total Contracts').font = total_font
            ws.cell(row=total_row, column=2, value=0).font = total_font
            ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=2)
            ws.row_dimensions[total_row].height = 35
            
            # Apply total row fill
            for col in range(1, 7):
                cell = ws.cell(row=total_row, column=col)
                cell.fill = total_fill
            
            # Set increased column widths
            for col_letter, width in [('A', 25), ('B', 70), ('C', 30), ('D', 25), ('E', 30), ('F', 35)]:
                ws.column_dimensions[col_letter].width = width
            
            # Add borders with custom blue color
            thin_border = Border(
                left=Side(style='thin', color="1E659E"),
                right=Side(style='thin', color="1E659E"),
                top=Side(style='thin', color="1E659E"),
                bottom=Side(style='thin', color="1E659E")
            )
            
            # Apply borders
            for row_num in range(1, total_row + 1):
                for cell in ws[row_num]:
                    cell.border = thin_border
            
            wb.save(output)
            output.seek(0)
            return send_file(
                output,
                download_name="empty_report.xlsx",
                as_attachment=True,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )